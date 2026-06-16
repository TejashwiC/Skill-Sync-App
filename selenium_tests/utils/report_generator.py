# =============================================================
# utils/report_generator.py — Excel report builder (openpyxl)
# =============================================================
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from config import REPORT_PREFIX, REPORT_DIR

# Column headers exactly as requested
HEADERS = [
    "S.No",
    "Test Case ID",
    "Module Name",
    "Test Scenario",
    "Expected Result",
    "Actual Result",
    "Status",
    "Remarks",
]


def generate_report(results: list[dict]) -> str:
    """
    Build a clean Excel report from a list of result dicts and return the file path.

    Each dict must contain:
        sno, tc_id, module, scenario, expected, actual, status, remarks
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "E2E Test Results"

    # ── Header row ────────────────────────────────────────
    ws.append(HEADERS)

    # Set column widths for readability
    col_widths = [6, 14, 20, 50, 45, 45, 10, 30]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # ── Data rows ──────────────────────────────────────────
    for r in results:
        ws.append([
            r.get("sno", ""),
            r.get("tc_id", ""),
            r.get("module", ""),
            r.get("scenario", ""),
            r.get("expected", ""),
            r.get("actual", ""),
            r.get("status", ""),
            r.get("remarks", ""),
        ])

    # ── Summary sheet ─────────────────────────────────────
    ws_summary = wb.create_sheet(title="Summary")
    total   = len(results)
    passed  = sum(1 for r in results if r.get("status", "").upper() == "PASS")
    failed  = sum(1 for r in results if r.get("status", "").upper() == "FAIL")
    skipped = sum(1 for r in results if r.get("status", "").upper() == "SKIP")
    pct     = round((passed / total * 100), 2) if total > 0 else 0.0

    ws_summary.append(["SkillSync E2E Test Summary"])
    ws_summary.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append([])
    ws_summary.append(["Metric", "Count"])
    ws_summary.append(["Total Test Cases", total])
    ws_summary.append(["Passed", passed])
    ws_summary.append(["Failed", failed])
    ws_summary.append(["Skipped", skipped])
    ws_summary.append(["Pass Percentage", f"{pct}%"])

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15

    # ── Save ──────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    filename  = f"{REPORT_PREFIX}_{timestamp}.xlsx"
    filepath  = os.path.join(REPORT_DIR, filename)
    wb.save(filepath)
    print(f"\n✅ Excel report saved → {filepath}")
    print(f"   Total: {total}  |  Passed: {passed}  |  Failed: {failed}  |  Skipped: {skipped}  |  Pass%: {pct}%\n")
    return filepath
