# =============================================================
# report_generator.py — Excel E2E Report for selenium_tests/
# Columns: S.No | Test Case ID | Module | Scenario |
#          Expected | Actual | Status | Remarks
# One file only, no colors/charts (clean format)
# =============================================================
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

MODULE_DISPLAY = {
    "test_01_auth":        "Authentication",
    "test_02_navigation":  "Navigation",
    "test_03_profile":     "Profile",
    "test_04_users":       "Users",
    "test_05_chat":        "Chat",
    "test_06_session":     "Session",
    "test_07_skilltest":   "Skill Test",
    "test_08_notes":       "Notes & PDFs",
    "test_09_settings":    "Settings",
    "test_10_security":    "Security",
}

HEADERS = [
    "S.No", "Test Case ID", "Module Name",
    "Test Scenario", "Expected Result", "Actual Result",
    "Status", "Remarks"
]

COL_WIDTHS = [6, 14, 20, 50, 40, 40, 10, 30]


def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_clr):
    return PatternFill(fill_type="solid", fgColor=hex_clr)


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_report(results: list, output_dir: str = ".") -> str:
    """
    Build Excel report with 2 sheets:
    1. E2E Test Results — one row per test
    2. Summary           — totals

    Deletes all old .xlsx files in output_dir first (one file rule).
    Returns file path.
    """
    os.makedirs(output_dir, exist_ok=True)

    # ── Guard ─────────────────────────────────────────────────────────────
    if not results:
        print("[Report] No results — Excel not generated.")
        return ""

    # ── Delete old reports ─────────────────────────────────────────────────
    for old in os.listdir(output_dir):
        if old.endswith(".xlsx"):
            try:
                os.remove(os.path.join(output_dir, old))
            except Exception:
                pass

    # ── Build workbook ─────────────────────────────────────────────────────
    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "E2E Test Results"
    ws2 = wb.create_sheet(title="Summary")

    _build_results_sheet(ws1, results)
    _build_summary_sheet(ws2, results)

    timestamp = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    filename  = f"E2E_Test_Report_SkillSync_{timestamp}.xlsx"
    filepath  = os.path.join(output_dir, filename)
    wb.save(filepath)

    total   = len(results)
    passed  = sum(1 for r in results if r["status"] == "PASS")
    failed  = sum(1 for r in results if r["status"] == "FAIL")
    skipped = sum(1 for r in results if r["status"] == "SKIP")
    pct     = round(passed / total * 100, 2) if total else 0

    print(f"\n{'='*60}")
    print(f"  Excel Report: {filepath}")
    print(f"  Total: {total}  Passed: {passed}  Failed: {failed}  Skipped: {skipped}  Pass%: {pct}%")
    print(f"{'='*60}\n")
    return filepath


def _build_results_sheet(ws, results):
    """Sheet 1: detailed row per test."""
    # Set column widths
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = "SkillSync — E2E Test Results"
    t.font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    t.fill      = _fill("1E3A5F")
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Header row
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = _font(bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("2E86AB")
        c.alignment = _center()
        c.border    = _thin()
    ws.row_dimensions[2].height = 22

    # Data rows
    for i, r in enumerate(results, start=1):
        row  = i + 2
        status = r.get("status", "SKIP")
        module = MODULE_DISPLAY.get(r.get("module", ""), r.get("module", ""))
        tc_id  = r.get("test_id", f"TC{i:03d}")
        name   = r.get("test_name", "")
        error  = r.get("error", "")

        row_data = [
            i,
            tc_id[:20],
            module,
            name,
            name,           # Expected = scenario description
            name if status == "PASS" else error,   # Actual
            status,
            error[:80] if status == "FAIL" else ("Skipped — credentials not set" if status == "SKIP" else ""),
        ]

        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.border    = _thin()
            c.font      = _font(size=9)
            c.alignment = _center() if col in (1, 7) else _left()

            # Status column colouring
            if col == 7:
                if status == "PASS":
                    c.fill = _fill("D6F5E3")
                    c.font = _font(bold=True, color="1A7A3C", size=9)
                elif status == "FAIL":
                    c.fill = _fill("FFE0E0")
                    c.font = _font(bold=True, color="C0392B", size=9)
                else:
                    c.fill = _fill("FFF8DC")
                    c.font = _font(bold=True, color="9A7D0A", size=9)
            elif col == 1:
                c.fill = _fill("ECF0F1")
            else:
                c.fill = _fill("F2F7FF" if i % 2 == 0 else "FFFFFF")

        ws.row_dimensions[row].height = 20

    # Footer totals
    total   = len(results)
    passed  = sum(1 for r in results if r["status"] == "PASS")
    failed  = sum(1 for r in results if r["status"] == "FAIL")
    skipped = sum(1 for r in results if r["status"] == "SKIP")
    footer_row = total + 3
    ws.merge_cells(f"A{footer_row}:H{footer_row}")
    f = ws[f"A{footer_row}"]
    f.value     = f"TOTAL: {total} tests  |  ✓ {passed} PASS  ✗ {failed} FAIL  ~ {skipped} SKIP"
    f.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    f.fill      = _fill("1E3A5F")
    f.alignment = _center()
    ws.row_dimensions[footer_row].height = 24


def _build_summary_sheet(ws, results):
    """Sheet 2: summary stats."""
    total   = len(results)
    passed  = sum(1 for r in results if r["status"] == "PASS")
    failed  = sum(1 for r in results if r["status"] == "FAIL")
    skipped = sum(1 for r in results if r["status"] == "SKIP")
    pct     = round(passed / total * 100, 2) if total else 0

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value     = "Test Summary"
    t.font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    t.fill      = _fill("1E3A5F")
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    rows = [
        ("Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Application", "SkillSync"),
        ("Framework", "Python + Selenium + pytest"),
        ("", ""),
        ("Total Test Cases", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Skipped", skipped),
        ("Pass Percentage", f"{pct}%"),
    ]
    for i, (label, value) in enumerate(rows, start=2):
        ws.row_dimensions[i].height = 22
        la = ws.cell(row=i, column=1, value=label)
        va = ws.cell(row=i, column=2, value=value)
        if label:
            la.font      = _font(bold=True, color="1E3A5F", size=10)
            la.fill      = _fill("EAF2FB")
            la.alignment = _left()
            la.border    = _thin()
            va.font      = _font(size=10)
            va.fill      = _fill("FDFEFE")
            va.alignment = _center()
            va.border    = _thin()
