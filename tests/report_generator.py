"""
report_generator.py — Generates an Excel E2E Test Report for SkillSync
Matches the format of: E2E_Test_Report_PancreaScan_2026-06-09T16-22-48.xlsx
"""
import sys
# Fix Windows console encoding for emoji/unicode
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except AttributeError:
        pass

import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList

# ─── Color Palette ─────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1E3A5F"   # Dark navy for main headers
CLR_HEADER_FG   = "FFFFFF"   # White text
CLR_PASS_BG     = "D6F5E3"   # Light green
CLR_PASS_FG     = "1A7A3C"   # Dark green
CLR_FAIL_BG     = "FFE0E0"   # Light red
CLR_FAIL_FG     = "C0392B"   # Dark red
CLR_SKIP_BG     = "FFF8DC"   # Light yellow
CLR_SKIP_FG     = "9A7D0A"   # Dark yellow
CLR_ROW_ALT     = "F2F7FF"   # Alternate row blue-tint
CLR_ROW_WHITE   = "FFFFFF"
CLR_TITLE_BG    = "2E86AB"   # Title section blue
CLR_SUMMARY_HDR = "4A90D9"   # Summary sub-headers

# Module color map (each test module gets a distinct color badge)
MODULE_COLORS = {
    "test_01_login":            ("D1F2EB", "1A5276"),
    "test_02_register":         ("FDEBD0", "A04000"),
    "test_03_forgot_password":  ("E8DAEF", "6C3483"),
    "test_04_dashboard_home":   ("D5F5E3", "145A32"),
    "test_05_profile":          ("D6EAF8", "1A5276"),
    "test_06_users":            ("FDFBD4", "7D6608"),
    "test_07_chat":             ("FADBD8", "922B21"),
    "test_08_session":          ("D5D8DC", "2E4053"),
    "test_09_skill_test":       ("D1F2EB", "117A65"),
    "test_10_notes":            ("FDEDEC", "C0392B"),
    "test_11_settings":         ("EBF5FB", "1B4F72"),
    # New modules
    "test_12_api_unit":         ("E8F8F5", "0E6655"),   # Teal — API Unit
    "test_13_validation":       ("FEF9E7", "7D6608"),   # Amber — Validation
    "test_14_vulnerability":    ("FDEDEC", "7B241C"),   # Red   — Security
}

TEST_TYPE_MAP = {
    "register":       "Validation",
    "login":          "Functional",
    "forgot":         "Validation",
    "home":           "Functional",
    "profile":        "Functional",
    "users":          "Functional",
    "chat":           "Functional",
    "session":        "Functional",
    "skill_test":     "Functional",
    "notes":          "Unit",
    "settings":       "Functional",
    "api_unit":       "API Unit",
    "validation":     "Validation",
    "vulnerability":  "Security",
}

PRIORITY_MAP = {
    "FAIL": "High",
    "PASS": "Medium",
    "SKIP": "Low",
}


def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _thick_border():
    thick = Side(style="medium", color="1E3A5F")
    return Border(left=thick, right=thick, top=thick, bottom=thick)


def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def _auto_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ─── Sheet 1: Summary Dashboard ────────────────────────────────────────────
def _build_summary_sheet(wb, results, timestamp):
    ws = wb.active
    ws.title = "📊 Summary"
    ws.sheet_view.showGridLines = False

    total   = len(results)
    passed  = sum(1 for r in results if r["status"] == "PASS")
    failed  = sum(1 for r in results if r["status"] == "FAIL")
    skipped = sum(1 for r in results if r["status"] == "SKIP")
    pass_rate = (passed / total * 100) if total > 0 else 0
    total_time = sum(r.get("duration", 0) for r in results)

    # Column widths
    _auto_col_widths(ws, {
        "A": 4, "B": 28, "C": 22, "D": 18, "E": 18,
        "F": 18, "G": 18, "H": 4
    })

    # Row heights
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 55
    ws.row_dimensions[3].height = 8

    # ── Title Banner ─────────────────────────────────────────────────────
    ws.merge_cells("B2:G2")
    title_cell = ws["B2"]
    title_cell.value = "🎯  SkillSync — E2E Test Execution Report"
    title_cell.fill = _fill(CLR_TITLE_BG)
    title_cell.font = Font(bold=True, color="FFFFFF", size=22, name="Calibri")
    title_cell.alignment = _center()
    title_cell.border = _thick_border()

    # ── Metadata block ───────────────────────────────────────────────────
    meta_rows = [
        ("Application",   "SkillSync — Skill Sharing Platform"),
        ("Test Date",     timestamp.strftime("%B %d, %Y")),
        ("Execution Time", timestamp.strftime("%I:%M:%S %p")),
        ("Environment",   "Chrome (Headless) / Firebase Live"),
        ("Test Framework", "Python + Selenium + pytest"),
        ("Report Version", "1.0"),
    ]
    start_row = 4
    for i, (label, value) in enumerate(meta_rows):
        row = start_row + i
        ws.row_dimensions[row].height = 22

        ws.merge_cells(f"B{row}:C{row}")
        lbl = ws[f"B{row}"]
        lbl.value = label
        lbl.font = _font(bold=True, color=CLR_HEADER_BG, size=11)
        lbl.fill = _fill("EAF2FB")
        lbl.alignment = _left()
        lbl.border = _thin_border()

        ws.merge_cells(f"D{row}:G{row}")
        val = ws[f"D{row}"]
        val.value = value
        val.font = _font(size=11)
        val.fill = _fill("FDFEFE")
        val.alignment = _left()
        val.border = _thin_border()

    # ── Spacer ────────────────────────────────────────────────────────────
    spacer_row = start_row + len(meta_rows) + 1
    ws.row_dimensions[spacer_row].height = 10

    # ── Summary Stats Cards ───────────────────────────────────────────────
    cards_row = spacer_row + 1
    ws.row_dimensions[cards_row].height = 16
    ws.row_dimensions[cards_row + 1].height = 38
    ws.row_dimensions[cards_row + 2].height = 28
    ws.row_dimensions[cards_row + 3].height = 10

    cards = [
        ("B", "C", "Total Tests",   str(total),          "2E86AB", "FFFFFF"),
        ("D", "D", "✅ Passed",      str(passed),         "27AE60", "FFFFFF"),
        ("E", "E", "❌ Failed",      str(failed),         "E74C3C", "FFFFFF"),
        ("F", "F", "⏭ Skipped",     str(skipped),        "F39C12", "FFFFFF"),
        ("G", "G", "Pass Rate",     f"{pass_rate:.1f}%", "8E44AD", "FFFFFF"),
    ]

    # Header row for cards
    ws.merge_cells(f"B{cards_row}:G{cards_row}")
    hdr = ws[f"B{cards_row}"]
    hdr.value = "TEST EXECUTION SUMMARY"
    hdr.fill = _fill(CLR_HEADER_BG)
    hdr.font = _font(bold=True, color="FFFFFF", size=12)
    hdr.alignment = _center()
    hdr.border = _thin_border()

    for start_col, end_col, label, value, bg, fg in cards:
        ws.merge_cells(f"{start_col}{cards_row+1}:{end_col}{cards_row+1}")
        ws.merge_cells(f"{start_col}{cards_row+2}:{end_col}{cards_row+2}")

        label_cell = ws[f"{start_col}{cards_row+1}"]
        label_cell.value = label
        label_cell.fill = _fill(bg)
        label_cell.font = _font(bold=True, color=fg, size=10)
        label_cell.alignment = _center()
        label_cell.border = _thin_border()

        value_cell = ws[f"{start_col}{cards_row+2}"]
        value_cell.value = value
        value_cell.fill = _fill(bg)
        value_cell.font = Font(bold=True, color=fg, size=20, name="Calibri")
        value_cell.alignment = _center()
        value_cell.border = _thin_border()

    # ── Additional stats ──────────────────────────────────────────────────
    extra_start = cards_row + 4
    extra_data = [
        ("Total Execution Time",  f"{total_time:.2f} seconds"),
        ("Avg. Test Duration",    f"{(total_time/total):.2f} sec" if total else "N/A"),
        ("Tests / Module",        str(len(set(r["module"] for r in results)))),
        ("Critical Failures",     str(failed)),
    ]
    for i, (label, value) in enumerate(extra_data):
        row = extra_start + i
        ws.row_dimensions[row].height = 22

        ws.merge_cells(f"B{row}:D{row}")
        lbl = ws[f"B{row}"]
        lbl.value = label
        lbl.font = _font(bold=True, color="2E4053", size=10)
        lbl.fill = _fill("EBF5FB")
        lbl.alignment = _left()
        lbl.border = _thin_border()

        ws.merge_cells(f"E{row}:G{row}")
        val = ws[f"E{row}"]
        val.value = value
        val.font = _font(size=10)
        val.fill = _fill("FDFEFE")
        val.alignment = _center()
        val.border = _thin_border()

    # ── Module Breakdown Table ────────────────────────────────────────────
    mod_start = extra_start + len(extra_data) + 2
    ws.row_dimensions[mod_start].height = 10

    header_row = mod_start + 1
    ws.row_dimensions[header_row].height = 28
    ws.merge_cells(f"B{header_row}:G{header_row}")
    mhdr = ws[f"B{header_row}"]
    mhdr.value = "MODULE BREAKDOWN"
    mhdr.fill = _fill(CLR_HEADER_BG)
    mhdr.font = _font(bold=True, color="FFFFFF", size=12)
    mhdr.alignment = _center()
    mhdr.border = _thin_border()

    col_headers = ["Module", "Total", "Pass", "Fail", "Skip", "Pass Rate"]
    col_cells   = ["B", "C", "D", "E", "F", "G"]
    col_row = header_row + 1
    ws.row_dimensions[col_row].height = 22
    for col, ch in zip(col_cells, col_headers):
        c = ws[f"{col}{col_row}"]
        c.value = ch
        c.fill = _fill(CLR_SUMMARY_HDR)
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = _center()
        c.border = _thin_border()

    # Group results by module
    modules = {}
    for r in results:
        m = r["module"]
        if m not in modules:
            modules[m] = {"total": 0, "pass": 0, "fail": 0, "skip": 0}
        modules[m]["total"] += 1
        if r["status"] == "PASS":
            modules[m]["pass"] += 1
        elif r["status"] == "FAIL":
            modules[m]["fail"] += 1
        else:
            modules[m]["skip"] += 1

    for j, (mod_name, counts) in enumerate(modules.items()):
        data_row = col_row + 1 + j
        ws.row_dimensions[data_row].height = 22
        pr = (counts["pass"] / counts["total"] * 100) if counts["total"] else 0
        bg = CLR_ROW_ALT if j % 2 == 0 else CLR_ROW_WHITE
        mod_colors = MODULE_COLORS.get(mod_name, ("F4F6F7", "2E4053"))

        row_data = [
            mod_name.replace("_", " ").title(),
            counts["total"],
            counts["pass"],
            counts["fail"],
            counts["skip"],
            f"{pr:.0f}%"
        ]
        for col, val in zip(col_cells, row_data):
            c = ws[f"{col}{data_row}"]
            c.value = val
            c.fill = _fill(mod_colors[0] if col == "B" else bg)
            c.font = _font(
                bold=(col == "B"),
                color=mod_colors[1] if col == "B" else "2E4053",
                size=10
            )
            c.alignment = _center() if col != "B" else _left()
            c.border = _thin_border()


# ─── Sheet 2: Full Test Results ────────────────────────────────────────────
def _build_results_sheet(wb, results, timestamp):
    ws = wb.create_sheet("📋 Test Results")
    ws.sheet_view.showGridLines = False

    # Column widths
    col_widths = {
        "A": 5,   "B": 6,   "C": 30,  "D": 25,
        "E": 42,  "F": 15,  "G": 10,  "H": 10,
        "I": 12,  "J": 45,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Freeze header
    ws.freeze_panes = "A4"

    # ── Title ─────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 8

    ws.merge_cells("A2:J2")
    title = ws["A2"]
    title.value = "SkillSync — Detailed E2E Test Results"
    title.fill = _fill(CLR_TITLE_BG)
    title.font = Font(bold=True, color="FFFFFF", size=18, name="Calibri")
    title.alignment = _center()
    title.border = _thick_border()

    # ── Column Headers ────────────────────────────────────────────────────
    headers = [
        ("#", "A"), ("Module", "B"), ("Module Name", "C"),
        ("Test Case ID", "D"), ("Test Case Name", "E"),
        ("Test Type", "F"), ("Priority", "G"), ("Status", "H"),
        ("Time (s)", "I"), ("Error / Notes", "J"),
    ]
    header_row = 4
    ws.row_dimensions[header_row].height = 30

    # Determine module display name from results
    MODULE_DISPLAY = {
        "test_01_login":            "01 - Login",
        "test_02_register":         "02 - Register",
        "test_03_forgot_password":  "03 - Forgot Pwd",
        "test_04_dashboard_home":   "04 - Home",
        "test_05_profile":          "05 - Profile",
        "test_06_users":            "06 - Users",
        "test_07_chat":             "07 - Chat",
        "test_08_session":          "08 - Session",
        "test_09_skill_test":       "09 - Skill Test",
        "test_10_notes":            "10 - Notes",
        "test_11_settings":         "11 - Settings",
        "test_12_api_unit":         "12 - API Unit",
        "test_13_validation":       "13 - Validation",
        "test_14_vulnerability":    "14 - Vulnerability",
    }

    for label, col in headers:
        c = ws[f"{col}{header_row}"]
        c.value = label
        c.fill = _fill(CLR_HEADER_BG)
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = _center()
        c.border = _thin_border()

    # ── Data Rows ─────────────────────────────────────────────────────────
    for i, r in enumerate(results):
        row = header_row + 1 + i
        ws.row_dimensions[row].height = 22

        status  = r.get("status", "SKIP")
        module  = r.get("module", "unknown")
        bg_alt  = CLR_ROW_ALT if i % 2 == 0 else CLR_ROW_WHITE

        # Determine test type from module name
        test_type = "Functional"
        for key, val in TEST_TYPE_MAP.items():
            if key in module:
                test_type = val
                break

        # Determine priority
        priority = PRIORITY_MAP.get(status, "Medium")

        # Status color
        if status == "PASS":
            status_bg, status_fg = CLR_PASS_BG, CLR_PASS_FG
        elif status == "FAIL":
            status_bg, status_fg = CLR_FAIL_BG, CLR_FAIL_FG
        else:
            status_bg, status_fg = CLR_SKIP_BG, CLR_SKIP_FG

        # Module color
        mod_colors = MODULE_COLORS.get(module, ("F4F6F7", "2E4053"))

        row_values = {
            "A": i + 1,
            "B": module.split("_")[1] if "_" in module else "",
            "C": MODULE_DISPLAY.get(module, module),
            "D": r.get("test_id", "")[:25],
            "E": r.get("test_name", ""),
            "F": test_type,
            "G": priority,
            "H": status,
            "I": r.get("duration", 0),
            "J": r.get("error", ""),
        }

        for col_letter, value in row_values.items():
            c = ws[f"{col_letter}{row}"]
            c.value = value
            c.border = _thin_border()

            if col_letter == "H":  # Status column
                c.fill = _fill(status_bg)
                c.font = _font(bold=True, color=status_fg, size=10)
                c.alignment = _center()
            elif col_letter in ("B", "C"):  # Module columns
                c.fill = _fill(mod_colors[0])
                c.font = _font(bold=(col_letter == "C"), color=mod_colors[1], size=9)
                c.alignment = _center() if col_letter == "B" else _left()
            elif col_letter == "A":  # Row number
                c.fill = _fill("ECF0F1")
                c.font = _font(color="7F8C8D", size=9)
                c.alignment = _center()
            elif col_letter == "J":  # Error notes
                c.fill = _fill(CLR_FAIL_BG if status == "FAIL" else bg_alt)
                c.font = _font(color=CLR_FAIL_FG if status == "FAIL" else "555555",
                               size=9, italic=(status == "FAIL"))
                c.alignment = _left()
            elif col_letter in ("F", "G", "I"):
                c.fill = _fill(bg_alt)
                c.font = _font(size=10)
                c.alignment = _center()
            else:
                c.fill = _fill(bg_alt)
                c.font = _font(size=10)
                c.alignment = _left()

    # ── Footer totals row ─────────────────────────────────────────────────
    total_row = header_row + 1 + len(results)
    ws.row_dimensions[total_row].height = 26
    ws.merge_cells(f"A{total_row}:G{total_row}")
    summary_cell = ws[f"A{total_row}"]
    summary_cell.value = f"TOTAL: {len(results)} tests | ✅ {sum(1 for r in results if r['status']=='PASS')} PASS  ❌ {sum(1 for r in results if r['status']=='FAIL')} FAIL  ⏭ {sum(1 for r in results if r['status']=='SKIP')} SKIP"
    summary_cell.fill = _fill(CLR_HEADER_BG)
    summary_cell.font = _font(bold=True, color="FFFFFF", size=11)
    summary_cell.alignment = _center()
    summary_cell.border = _thick_border()


# ─── Sheet 3: Charts ───────────────────────────────────────────────────────
def _build_charts_sheet(wb, results):
    ws = wb.create_sheet("📈 Charts")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Status"
    ws["B1"] = "Count"
    statuses = {"PASS": 0, "FAIL": 0, "SKIP": 0}
    for r in results:
        statuses[r.get("status", "SKIP")] += 1

    ws["A2"] = "PASS";  ws["B2"] = statuses["PASS"]
    ws["A3"] = "FAIL";  ws["B3"] = statuses["FAIL"]
    ws["A4"] = "SKIP";  ws["B4"] = statuses["SKIP"]

    # Pie chart
    pie = PieChart()
    pie.title = "Test Results Distribution"
    pie.style = 10
    labels = Reference(ws, min_col=1, min_row=2, max_row=4)
    data   = Reference(ws, min_col=2, min_row=1, max_row=4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width  = 18
    pie.height = 14
    ws.add_chart(pie, "D2")

    # Module bar data
    ws["A7"]  = "Module"
    ws["B7"]  = "Pass"
    ws["C7"]  = "Fail"

    modules = {}
    for r in results:
        m = r["module"]
        if m not in modules:
            modules[m] = {"pass": 0, "fail": 0}
        if r["status"] == "PASS":
            modules[m]["pass"] += 1
        elif r["status"] == "FAIL":
            modules[m]["fail"] += 1

    for j, (m, counts) in enumerate(modules.items()):
        row = 8 + j
        ws.cell(row=row, column=1).value = m.replace("_", " ").title()
        ws.cell(row=row, column=2).value = counts["pass"]
        ws.cell(row=row, column=3).value = counts["fail"]

    # Bar chart
    bar = BarChart()
    bar.type   = "col"
    bar.title  = "Pass/Fail by Module"
    bar.style  = 10
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Module"

    data_ref   = Reference(ws, min_col=2, max_col=3, min_row=7, max_row=7 + len(modules))
    cats_ref   = Reference(ws, min_col=1, min_row=8, max_row=7 + len(modules))
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.shape  = 4
    bar.width  = 30
    bar.height = 18
    ws.add_chart(bar, "D20")


# ─── Main entry point ────────────────────────────────────────────────────────
def generate_excel_report(results, output_dir=None):
    """
    Generate the Excel report and save it.
    Called from conftest.py after all tests complete.

    Key behaviours:
    - Deletes ALL existing .xlsx files in output_dir before saving (keeps only 1 file).
    - Skips saving entirely if results list is empty (prevents 0-row files).
    """
    if output_dir is None:
        # Save to tests/reports/ relative to this file
        output_dir = os.path.join(os.path.dirname(__file__), "reports")

    os.makedirs(output_dir, exist_ok=True)

    # ── Guard: never write an empty report ───────────────────────────────────
    if not results:
        print("\n[REPORT] No test results collected — Excel report NOT generated.\n")
        return None

    # ── Delete all previous .xlsx files so only ONE report exists ────────────
    for old_file in os.listdir(output_dir):
        if old_file.endswith(".xlsx"):
            try:
                os.remove(os.path.join(output_dir, old_file))
            except Exception:
                pass  # ignore locked files

    timestamp = datetime.datetime.now()
    filename  = f"E2E_Test_Report_SkillSync_{timestamp.strftime('%Y-%m-%dT%H-%M-%S')}.xlsx"
    filepath  = os.path.join(output_dir, filename)

    wb = Workbook()

    _build_summary_sheet(wb, results, timestamp)
    _build_results_sheet(wb, results, timestamp)
    _build_charts_sheet(wb, results)

    wb.save(filepath)
    passed  = sum(1 for r in results if r['status'] == 'PASS')
    failed  = sum(1 for r in results if r['status'] == 'FAIL')
    skipped = sum(1 for r in results if r['status'] == 'SKIP')
    sep = '=' * 60
    print(f"\n{sep}")
    print(f"  [REPORT] Excel Report Generated:")
    print(f"  {filepath}")
    print(f"  Total Tests : {len(results)}")
    print(f"  Passed      : {passed}")
    print(f"  Failed      : {failed}")
    print(f"  Skipped     : {skipped}")
    print(f"{sep}\n")

    return filepath


# ─── Standalone run for testing the report generator itself ─────────────────
if __name__ == "__main__":
    # Build 174 sample results — one for each real test
    sample_data = [
        # Module 01 - Login (16)
        ("test_01_login", "LOG_001", "Page Loads With Title"),
        ("test_01_login", "LOG_002", "Logo Visible"),
        ("test_01_login", "LOG_003", "Subtitle Shown"),
        ("test_01_login", "LOG_004", "Email Field Present"),
        ("test_01_login", "LOG_005", "Password Field Present"),
        ("test_01_login", "LOG_006", "Login Button Present"),
        ("test_01_login", "LOG_007", "Forgot Password Link Present"),
        ("test_01_login", "LOG_008", "Forgot Password Link Navigates"),
        ("test_01_login", "LOG_009", "Register Link Present"),
        ("test_01_login", "LOG_010", "Register Link Navigates"),
        ("test_01_login", "LOG_011", "Empty Email HTML5 Validation"),
        ("test_01_login", "LOG_012", "Invalid Email Format"),
        ("test_01_login", "LOG_013", "Wrong Credentials Alert"),
        ("test_01_login", "LOG_014", "Inputs Accept Text"),
        ("test_01_login", "LOG_015", "Inputs Clearable"),
        ("test_01_login", "LOG_016", "Valid Login Redirects"),
        # Module 02 - Register (14)
        ("test_02_register", "REG_001", "Page Loads"),
        ("test_02_register", "REG_002", "Logo Displayed"),
        ("test_02_register", "REG_003", "Subtitle Create Account"),
        ("test_02_register", "REG_004", "Name Field Present"),
        ("test_02_register", "REG_005", "Email Field Type"),
        ("test_02_register", "REG_006", "Password Field Type"),
        ("test_02_register", "REG_007", "Register Button Visible"),
        ("test_02_register", "REG_008", "Login Link Present"),
        ("test_02_register", "REG_009", "Login Link Navigates"),
        ("test_02_register", "REG_010", "Empty Name HTML5 Validation"),
        ("test_02_register", "REG_011", "Invalid Email Format"),
        ("test_02_register", "REG_012", "All Inputs Accept Text"),
        ("test_02_register", "REG_013", "CSS Stylesheet Applied"),
        ("test_02_register", "REG_014", "Page Has Three Inputs"),
        # Module 03 - Forgot Password (13)
        ("test_03_forgot_password", "FP_001", "Page Loads"),
        ("test_03_forgot_password", "FP_002", "Logo Visible"),
        ("test_03_forgot_password", "FP_003", "Subtitle Shown"),
        ("test_03_forgot_password", "FP_004", "Email Field Present"),
        ("test_03_forgot_password", "FP_005", "Send Reset Button Present"),
        ("test_03_forgot_password", "FP_006", "Success Msg Hidden Default"),
        ("test_03_forgot_password", "FP_007", "Error Msg Hidden Default"),
        ("test_03_forgot_password", "FP_008", "Empty Email Shows Error"),
        ("test_03_forgot_password", "FP_009", "Invalid Email Format"),
        ("test_03_forgot_password", "FP_010", "Back To Login Link"),
        ("test_03_forgot_password", "FP_011", "Back To Login Navigates"),
        ("test_03_forgot_password", "FP_012", "Email Field Accepts Text"),
        ("test_03_forgot_password", "FP_013", "Page Has One Input"),
        # Module 04 - Dashboard Home (20)
        ("test_04_dashboard_home", "HOME_001", "Dashboard Loads"),
        ("test_04_dashboard_home", "HOME_002", "Sidebar Present"),
        ("test_04_dashboard_home", "HOME_003", "Logo In Sidebar"),
        ("test_04_dashboard_home", "HOME_004", "Welcome Text Shown"),
        ("test_04_dashboard_home", "HOME_005", "Home Section Visible"),
        ("test_04_dashboard_home", "HOME_006", "Skills Stat Card"),
        ("test_04_dashboard_home", "HOME_007", "Credits Stat Card"),
        ("test_04_dashboard_home", "HOME_008", "Sessions Stat Card"),
        ("test_04_dashboard_home", "HOME_009", "Tests Created Card"),
        ("test_04_dashboard_home", "HOME_010", "Tests Completed Card"),
        ("test_04_dashboard_home", "HOME_011", "Dashboard Grid Rendered"),
        ("test_04_dashboard_home", "HOME_012", "Nav Profile"),
        ("test_04_dashboard_home", "HOME_013", "Nav Users"),
        ("test_04_dashboard_home", "HOME_014", "Nav Chat"),
        ("test_04_dashboard_home", "HOME_015", "Nav Session"),
        ("test_04_dashboard_home", "HOME_016", "Nav Skill Test"),
        ("test_04_dashboard_home", "HOME_017", "Nav Notes"),
        ("test_04_dashboard_home", "HOME_018", "Nav Settings"),
        ("test_04_dashboard_home", "HOME_019", "Nav Home Returns"),
        ("test_04_dashboard_home", "HOME_020", "Logout Button Present"),
        # Module 05 - Profile (16)
        ("test_05_profile", "PROF_001", "Section Loads"),
        ("test_05_profile", "PROF_002", "Dashboard Grid Cards"),
        ("test_05_profile", "PROF_003", "View Profile Opens"),
        ("test_05_profile", "PROF_004", "View Profile Fields"),
        ("test_05_profile", "PROF_005", "View Profile Back Btn"),
        ("test_05_profile", "PROF_006", "Edit Profile Opens"),
        ("test_05_profile", "PROF_007", "Edit Profile All Fields"),
        ("test_05_profile", "PROF_008", "Save Button Present"),
        ("test_05_profile", "PROF_009", "Add Skill Screen Opens"),
        ("test_05_profile", "PROF_010", "Add Skill Input Present"),
        ("test_05_profile", "PROF_011", "Add Skill Empty Validation"),
        ("test_05_profile", "PROF_012", "Edit Skill Screen Opens"),
        ("test_05_profile", "PROF_013", "Delete Skill Screen Opens"),
        ("test_05_profile", "PROF_014", "Skills Overview Opens"),
        ("test_05_profile", "PROF_015", "Recommendations Opens"),
        ("test_05_profile", "PROF_016", "Upload Photo Screen Opens"),
        # Module 06 - Users (14)
        ("test_06_users", "USR_001", "Section Loads"),
        ("test_06_users", "USR_002", "Dashboard Heading"),
        ("test_06_users", "USR_003", "Cards Count"),
        ("test_06_users", "USR_004", "Users List Opens"),
        ("test_06_users", "USR_005", "Users List Back Btn"),
        ("test_06_users", "USR_006", "Followers Screen Opens"),
        ("test_06_users", "USR_007", "Following Screen Opens"),
        ("test_06_users", "USR_008", "Search Users Screen Opens"),
        ("test_06_users", "USR_009", "Search Input Present"),
        ("test_06_users", "USR_010", "Search With Query Shows Results"),
        ("test_06_users", "USR_011", "Search Clear Resets Results"),
        ("test_06_users", "USR_012", "Suggested Users Opens"),
        ("test_06_users", "USR_013", "Top Mentors Screen Opens"),
        ("test_06_users", "USR_014", "Users Section Heading Visible"),
        # Module 07 - Chat (17)
        ("test_07_chat", "CHAT_001", "Section Loads"),
        ("test_07_chat", "CHAT_002", "Dashboard Heading"),
        ("test_07_chat", "CHAT_003", "Cards Count"),
        ("test_07_chat", "CHAT_004", "Main Chat Opens"),
        ("test_07_chat", "CHAT_005", "Message Input Present"),
        ("test_07_chat", "CHAT_006", "Send Button Present"),
        ("test_07_chat", "CHAT_007", "Chat User List Present"),
        ("test_07_chat", "CHAT_008", "Chat Box Present"),
        ("test_07_chat", "CHAT_009", "Send Empty No Crash"),
        ("test_07_chat", "CHAT_010", "Chat Requests Opens"),
        ("test_07_chat", "CHAT_011", "Inbox Opens"),
        ("test_07_chat", "CHAT_012", "Inbox Container Present"),
        ("test_07_chat", "CHAT_013", "Chat Settings Opens"),
        ("test_07_chat", "CHAT_014", "Notification Toggles Present"),
        ("test_07_chat", "CHAT_015", "Blocked Users Opens"),
        ("test_07_chat", "CHAT_016", "Blocked Users List Present"),
        ("test_07_chat", "CHAT_017", "Group Chat Opens"),
        # Module 08 - Session (21)
        ("test_08_session", "SESS_001", "Section Loads"),
        ("test_08_session", "SESS_002", "Heading Text"),
        ("test_08_session", "SESS_003", "Cards Count"),
        ("test_08_session", "SESS_004", "Start Session Screen"),
        ("test_08_session", "SESS_005", "Session Name Field"),
        ("test_08_session", "SESS_006", "Platform Dropdown"),
        ("test_08_session", "SESS_007", "Meeting Link URL Field"),
        ("test_08_session", "SESS_008", "Empty Name Alert"),
        ("test_08_session", "SESS_009", "Empty Platform Alert"),
        ("test_08_session", "SESS_010", "Invalid Link Alert"),
        ("test_08_session", "SESS_011", "Platform Links Present"),
        ("test_08_session", "SESS_012", "Join Session Screen"),
        ("test_08_session", "SESS_013", "Join Code MaxLength"),
        ("test_08_session", "SESS_014", "Short Code Alert"),
        ("test_08_session", "SESS_015", "Invalid Code Shows Result"),
        ("test_08_session", "SESS_016", "Live Sessions Screen"),
        ("test_08_session", "SESS_017", "Session History Screen"),
        ("test_08_session", "SESS_018", "Feedback Screen Fields"),
        ("test_08_session", "SESS_019", "Feedback Empty Alert"),
        ("test_08_session", "SESS_020", "Ratings Screen Stars"),
        ("test_08_session", "SESS_021", "Star Rating Clickable"),
        # Module 09 - Skill Test (18)
        ("test_09_skill_test", "TST_001", "Section Loads"),
        ("test_09_skill_test", "TST_002", "Heading Text"),
        ("test_09_skill_test", "TST_003", "Cards Count"),
        ("test_09_skill_test", "TST_004", "Create Test Screen"),
        ("test_09_skill_test", "TST_005", "Create Test Title Field"),
        ("test_09_skill_test", "TST_006", "Create Test Skill Field"),
        ("test_09_skill_test", "TST_007", "Create Test Credits Number"),
        ("test_09_skill_test", "TST_008", "Empty Title Alert"),
        ("test_09_skill_test", "TST_009", "Empty Skill Alert"),
        ("test_09_skill_test", "TST_010", "Add Questions Screen"),
        ("test_09_skill_test", "TST_011", "Question And Options Fields"),
        ("test_09_skill_test", "TST_012", "Correct Answer Dropdown"),
        ("test_09_skill_test", "TST_013", "Option Populates Dropdown"),
        ("test_09_skill_test", "TST_014", "My Questions Screen"),
        ("test_09_skill_test", "TST_015", "Attend Test Screen"),
        ("test_09_skill_test", "TST_016", "Available Tests Container"),
        ("test_09_skill_test", "TST_017", "My Results Screen"),
        ("test_09_skill_test", "TST_018", "Leaderboard Screen"),
        # Module 10 - Notes (11)
        ("test_10_notes", "NOTES_001", "Section Loads"),
        ("test_10_notes", "NOTES_002", "Heading Text"),
        ("test_10_notes", "NOTES_003", "Dashboard Cards"),
        ("test_10_notes", "NOTES_004", "Upload Notes Screen"),
        ("test_10_notes", "NOTES_005", "PDF File Input Accept"),
        ("test_10_notes", "NOTES_006", "Upload Button Present"),
        ("test_10_notes", "NOTES_007", "Progress Bar Hidden Default"),
        ("test_10_notes", "NOTES_008", "View Notes Screen"),
        ("test_10_notes", "NOTES_009", "PDF List Container"),
        ("test_10_notes", "NOTES_010", "View Notes Back Btn"),
        ("test_10_notes", "NOTES_011", "Upload Notes Back Btn"),
        # Module 11 - Settings (14)
        ("test_11_settings", "SET_001", "Section Loads"),
        ("test_11_settings", "SET_002", "Heading Text"),
        ("test_11_settings", "SET_003", "Cards Count"),
        ("test_11_settings", "SET_004", "Account Settings Opens"),
        ("test_11_settings", "SET_005", "Account Settings Inputs"),
        ("test_11_settings", "SET_006", "Notification Settings Opens"),
        ("test_11_settings", "SET_007", "Follow Notification Toggles"),
        ("test_11_settings", "SET_008", "Sound Notification Toggles"),
        ("test_11_settings", "SET_009", "Email Notification Toggles"),
        ("test_11_settings", "SET_010", "Session Alert Toggles"),
        ("test_11_settings", "SET_011", "Security Settings Opens"),
        ("test_11_settings", "SET_012", "Password Settings Opens"),
        ("test_11_settings", "SET_013", "Password Three Fields"),
        ("test_11_settings", "SET_014", "All Back Buttons Work"),
    ]

    import random
    sample_results = []
    for module, tid, tname in sample_data:
        r = random.random()
        status   = "PASS" if r > 0.15 else ("FAIL" if r > 0.05 else "SKIP")
        duration = round(random.uniform(0.3, 4.5), 2)
        error    = "AssertionError: Element not found" if status == "FAIL" else ""
        sample_results.append({
            "module":    module,
            "test_id":   tid,
            "test_name": tname,
            "status":    status,
            "duration":  duration,
            "error":     error,
        })

    path = generate_excel_report(sample_results)
    print(f"Sample report with {len(sample_results)} tests: {path}")
